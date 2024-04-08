# -*- coding: windows-1250 -*-

import openpyxl
import matplotlib.pyplot as plt
import numpy as np


def create_blank_excel(constituencies, parties):
    # Create a new blank Excel workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Column names (constituencies)
    for col, constituency in enumerate(constituencies, start=2):
        sheet.cell(row=1, column=col, value=constituency)

    # Row names (political parties)
    for row, party in enumerate(parties, start=2):
        sheet.cell(row=row, column=1, value=party)

    return wb, sheet


def populate_party_votes(workbook, sheet, party_votes):
    for row, party in enumerate(party_votes, start=2):
        for col, votes in enumerate(party_votes[party], start=2):
            sheet.cell(row=row, column=col, value=votes)
    workbook.save("party_errors.xlsx")


def populate_excel_workbook(constituencies_data):
    party_errors = {party: [] for party in parties}
    for party in parties:
        for c in constituencies_data.keys():
            for p in constituencies_data[c]:
                if p[0] == party:
                    party_errors[party].append(float(p[2]) / float(p[1]))

    # Create blank Excel file and populate with party votes
    workbook, sheet = create_blank_excel(constituencies_data, parties)
    populate_party_votes(workbook, sheet, party_errors)


def show_poll_data(constituencies_data):
    for i, (constituency, constituency_data) in enumerate(constituencies_data.items(), start=1):
        party_names = [party[0] for party in constituency_data]
        predicted_percentages = [float(party[1]) for party in constituency_data]
        actual_percentages = [float(party[2]) for party in constituency_data]

        # Create figure and axes
        fig, ax = plt.subplots()

        # Set positions for bars
        bar_positions = np.arange(len(party_names))
        width = 0.35

        # Plot predicted percentages
        ax.bar(bar_positions - width / 2, predicted_percentages, width, label='Predviđeni postotak')

        # Plot actual percentages
        ax.bar(bar_positions + width / 2, actual_percentages, width, label='Stvarni postotak')

        # Set x-axis ticks and labels
        ax.set_xticks(bar_positions)
        ax.set_xticklabels(party_names)

        # Set labels and title
        ax.set_xlabel('Stranka')
        ax.set_ylabel('Postotak glasova (%)')
        ax.set_title(f'{i}. izborna jedinica')

        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45, ha='right')

        # Add legend
        ax.legend()

        # Show plot
        plt.tight_layout()  # Adjust layout to prevent overlapping
        plt.show()


constituencies_names = [str(i) + ". izborna jedinica" for i in range(1, 11)]

parties = ["HDZ", "RESTART", "MOŽEMO", "MOST", "DP"]


with open("poll_errors.txt", "r") as file:
    text = file.read().split("\n")

    constituencies = {}
    cur_const = ""
    for line in text:
        if not line: continue
        if len(line) <= 4:
            constituencies[line] = []
            cur_const = line
        else:
            line2 = line.split(" ")
            party = line2[0]
            predicted, actual = line2[1].split(",")
            constituencies[cur_const].append((party, predicted, actual))


show_poll_data(constituencies)
